import pandas as pd
import openpyxl
import os
from fuzzer import cirq_qiskit_diff_test

transpile_fail, suspicious, success = cirq_qiskit_diff_test()

dataframe = pd.DataFrame(suspicious, columns=["cirq", "qiskit", "cirq_prime", "qiskit_prime"])
os.chdir("/Users/ihayeong/Downloads")
dataframe.to_excel("suspicious.xlsx", index=False)

def update_excel_file(filename, suspicious):
    dataframe = pd.DataFrame(suspicious, columns=["cirq", "qiskit", "cirq_prime", "qiskit_prime"])
    
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # os.chdir("/Users/ihayeong/Downloads")
        dataframe.to_excel("suspicious.xlsx", index=False)
        return
    
    sheet = workbook.active
    next_row = sheet.max_row + 1 # next available row

    for row_index, item in enumerate(dataframe, start=next_row): # append data
        sheet.cell(row=row_index, column=1, value=item)

    workbook.save(filename)
    workbook.close()